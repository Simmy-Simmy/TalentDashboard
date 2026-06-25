from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\Admin\Desktop\talent-dashboard")
DOWNLOADS = Path(r"C:\Users\Admin\Downloads")
WORKBOOK = DOWNLOADS / "talent_dashboard_clean_output.xlsx"
FEEDBACK_CSV = DOWNLOADS / "export_feedback.csv"
CALIBRATION_CSV = DOWNLOADS / "export_interview_calibration.csv"

DATA_DIR = ROOT / "data"
OUTPUT_JS = ROOT / "data-loader.js"

STAGE_ORDER = {
    "HR Interview": 0,
    "1st Interview": 1,
    "2nd Interview": 2,
    "3rd Interview": 3,
    "Unknown": 4,
}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def short_note(value: str, limit: int = 60) -> str:
    text = clean(value)
    if not text:
        return "Not captured in ATS."
    if len(text) <= limit:
        return text
    cut = text[: limit - 1].rstrip()
    if " " in cut:
        cut = cut.rsplit(" ", 1)[0]
    return f"{cut}..."


def first_sentence(value: str, limit: int = 120) -> str:
    text = clean(value)
    if not text:
        return "Not captured in ATS."
    parts = re.split(r"(?<=[.!?])\s+", text)
    sentence = parts[0] if parts else text
    if len(sentence) <= limit:
        return sentence
    cut = sentence[: limit - 1].rstrip()
    if " " in cut:
        cut = cut.rsplit(" ", 1)[0]
    return f"{cut}..."


def slug(text: str) -> str:
    text = clean(text).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "item"


def dedupe_headers(headers):
    counts = {}
    result = []
    for index, header in enumerate(headers):
        key = clean(header) or f"col_{index + 1}"
        counts[key] = counts.get(key, 0) + 1
        result.append(key if counts[key] == 1 else f"{key}__{counts[key]}")
    return result


def parse_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if len(rows) < 3:
        return []
    headers = dedupe_headers(rows[1])
    data = []
    for row in rows[2:]:
        if not any(clean(cell) for cell in row):
            continue
        padded = row + [""] * (len(headers) - len(row))
        item = {headers[i]: padded[i] if i < len(padded) else "" for i in range(len(headers))}
        data.append(item)
    return data


def parse_number_from_text(text):
    text = clean(text)
    if not text:
        return None
    numbers = [float(match) for match in re.findall(r"(?<!\d)(\d(?:\.\d)?)", text)]
    if not numbers:
        return None
    return round(sum(numbers) / len(numbers), 2)


def normalize_score(value):
    if value in (None, ""):
        return None
    text = clean(value)
    numbers = [float(match) for match in re.findall(r"(?<!\d)(\d(?:\.\d)?)", text)]
    if not numbers:
        return None
    return round(sum(numbers) / len(numbers), 2)


def yes_no(value):
    text = clean(value).lower()
    if not text:
        return "Not captured in ATS."
    if text in {"yes", "y", "true"}:
        return "Yes"
    if text in {"no", "n", "false"}:
        return "No"
    return clean(value)


def stage_rank(stage):
    return STAGE_ORDER.get(clean(stage), 99)


def sort_candidates(rows):
    return sorted(
        rows,
        key=lambda row: (
            clean(row["office_location"]) == "Unknown",
            clean(row["office_location"]),
            stage_rank(row["stage"]),
            -float(row["interview_score"]) if row["interview_score"] is not None else 99,
            clean(row["candidate_name"]),
        ),
    )


def read_workbook():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Candidates"]
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(record)
    funnel_ws = wb["Funnel Summary"]
    funnel_rows = [tuple(row) for row in funnel_ws.iter_rows(values_only=True)]
    return rows, funnel_rows


def build_candidate_rows(candidates_raw):
    candidates = []
    for row in candidates_raw:
        candidate_name = clean(row["candidate_name"])
        role = clean(row["role"])
        office = clean(row["office_location"]) or "Not captured in ATS."
        status = clean(row["status"]) or "Not captured in ATS."
        source = clean(row["source"]) or "Not captured in ATS."
        stage = clean(row["stage"]) or "Not captured in ATS."
        risk = clean(row["risk_level"]) or "Not captured in ATS."
        note = clean(row["next_step"]) or "Not captured in ATS."
        score = normalize_score(row["interview_score"])
        candidates.append(
            {
                "id": slug(f"{candidate_name}-{role}-{office}"),
                "candidate_name": candidate_name,
                "role": role,
                "office_location": office,
                "linkedin_url": None,
                "stage": stage,
                "status": status,
                "source": source,
                "interview_score": score,
                "risk_level": risk,
                "next_step": note,
                "last_updated": clean(row["last_updated"]) or "Not captured in ATS.",
            }
        )
    return sort_candidates(candidates)


def build_funnel_summary(funnel_rows):
    buckets = defaultdict(list)
    for row in funnel_rows[1:]:
        funnel_type, value, count = row
        buckets[clean(funnel_type)].append((clean(value), int(count or 0)))

    stage_order = ["HR Interview", "1st Interview", "2nd Interview", "3rd Interview", "Unknown"]
    stage_counts = {value: count for value, count in buckets.get("stage", [])}
    office_counts = {value: count for value, count in buckets.get("office_location", [])}
    status_counts = {value: count for value, count in buckets.get("status", [])}
    role_counts = {value: count for value, count in buckets.get("role", [])}
    risk_counts = {value: count for value, count in buckets.get("risk_level", [])}

    return {
        "stage_order": stage_order,
        "stage_counts": stage_counts,
        "office_counts": office_counts,
        "status_counts": status_counts,
        "role_counts": role_counts,
        "risk_counts": risk_counts,
    }


def parse_feedback_exports(path: Path, name_field: str):
    rows = parse_csv(path)
    records = []
    for row in rows:
        name = clean(row.get(name_field))
        if not name:
            continue
        records.append(row)
    return records


def load_json_rows(path: Path):
    if not path.exists():
        return []
    for encoding in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            return json.loads(path.read_text(encoding=encoding))
        except Exception:
            continue
    raise ValueError(f"Unable to read JSON rows from {path}")


def latest_detail(records, name_field):
    latest = {}
    for row in records:
        name = clean(row.get(name_field))
        if not name:
            continue
        date_text = clean(row.get("Interview Date")) or clean(row.get("Interview Date TZ")) or clean(row.get("Form Created at (UTC)"))
        try:
            parsed = datetime.fromisoformat(date_text.replace("Z", ""))
        except Exception:
            parsed = datetime.min
        current = latest.get(name)
        if not current or parsed >= current["_parsed"]:
            latest[name] = {"_parsed": parsed, **row}
    for name, row in list(latest.items()):
        row.pop("_parsed", None)
    return latest


def build_scorecard_from_row(row, source_kind):
    mapping = [
        ("consulting mindset", [
            "Consulting",
            "Basic Consulting Skills",
            "Structured Thinking & Use of Frameworks",
            "Consulting Skills: Relevant Client Insights",
            "Structured Thinking & \u000bUse of Frameworks",
        ]),
        ("problem solving", [
            "Business judgement",
            "Structured problem solving",
            "Information Synthesis",
            "Effectiveness & Impact",
        ]),
        ("client presence", [
            "Communication Skills",
            "Communication Skills: Oral Presentation of Content",
            "Client leadership under pressure",
            "Consulting Skills: Relevant Client Insights",
        ]),
        ("adaptability", [
            "Personal Attributes",
            "Effectiveness & Impact",
            "On-Site Work",
            "Willingness to Travel",
        ]),
        ("leadership", [
            "Leadership and Cultural Influence",
            "Leadership",
            "Strategic Networking and Client Engagement",
        ]),
        ("communication", [
            "Written Clarity",
            "Communication Skills",
            "Communication Skills: Oral Presentation of Content",
            "Structured communication",
        ]),
        ("commercial awareness", [
            "Business Development Skills",
            "Strategic Networking and Client Engagement",
            "Client Portfolio Expansion",
            "Business judgement",
        ]),
    ]

    scorecard = []
    for label, fields in mapping:
        scores = []
        notes = []
        for field in fields:
            value = row.get(field)
            if value in (None, ""):
                continue
            score = parse_number_from_text(value)
            if score is not None:
                scores.append(score)
            notes.append(clean(value))
        if scores:
            score_value = round(sum(scores) / len(scores))
            score_value = min(4, max(1, int(score_value)))
            note = short_note(" | ".join(notes), 58)
        else:
            score_value = None
            note = "Not captured in ATS."
        scorecard.append({"criteria": label, "score": score_value, "note": note, "source": source_kind})
    return scorecard


def build_detail_entry(candidate, feedback_map, calibration_map):
    name = candidate["candidate_name"]
    role = candidate["role"]
    office = candidate["office_location"]

    feedback = feedback_map.get(name)
    calibration = calibration_map.get(name)
    latest = {**(feedback or {}), **(calibration or {})}

    stage = clean(latest.get("Interview Stage")) or candidate["stage"]
    interview_date = clean(latest.get("Interview Date")) or clean(latest.get("Interview Date TZ")) or candidate["last_updated"]
    interviewer = clean(latest.get("Interviewer")) or clean(latest.get("Interviewer Name")) or "Not captured in ATS."
    overall_score = normalize_score(latest.get("Overall Score")) or normalize_score(latest.get("Overall Rating")) or candidate["interview_score"]
    recommendation = yes_no(latest.get("Do you recommend this applicant for next steps?"))
    red_flag = yes_no(latest.get("Red Flag Assessment (select one)"))
    trust = yes_no(latest.get("(SC & above) Would you trust sending them alone to the client, and would they be welcome on your case?"))
    what_test_next = clean(latest.get("What to test in next interview?")) or candidate["next_step"] or "Not captured in ATS."
    summary = first_sentence(latest.get("Overall Comments") or candidate["next_step"])
    strengths = summary
    concerns = clean(latest.get("Red Flag Assessment (select one)")) or clean(candidate["risk_level"])
    if concerns in {"", "No", "Not captured in ATS."}:
        concerns = "Not captured in ATS."
    stakeholder_summary = summary
    level = clean(latest.get("Which level is he/she more suited for or welcome on your team at?")) or "Not captured in ATS."
    latest_ref = clean(latest.get("Form Name") or latest.get("Feedback Form")) or "Not captured in ATS."
    if latest_ref != "Not captured in ATS.":
        latest_ref = f"{latest_ref} / {clean(interviewer)}"

    if latest:
        scorecard = build_scorecard_from_row(latest, "calibration" if calibration else "feedback")
    else:
        scorecard = [
            {"criteria": label, "score": None, "note": "Not captured in ATS.", "source": "none"}
            for label in [
                "consulting mindset",
                "problem solving",
                "client presence",
                "adaptability",
                "leadership",
                "communication",
                "commercial awareness",
            ]
        ]

    return {
        "id": candidate["id"],
        "candidate_name": name,
        "role": role,
        "office_location": office,
        "current_stage": stage,
        "company": "ADAPTOVATE",
        "interview_date": interview_date or "Not captured in ATS.",
        "latest_feedback_reference": latest_ref,
        "overall_rating": overall_score,
        "recommendation": recommendation or "Not captured in ATS.",
        "red_flag_status": red_flag or "Not captured in ATS.",
        "strengths": strengths,
        "concerns": concerns,
        "what_to_test_next": what_test_next,
        "stakeholder_summary": stakeholder_summary,
        "suggested_level": level,
        "client_trust": trust,
        "scorecard": scorecard,
        "next_step": candidate["next_step"],
        "status": candidate["status"],
        "risk_level": candidate["risk_level"],
        "source": candidate["source"],
        "last_updated": candidate["last_updated"],
        "notes": clean(latest.get("Overall Comments")) or "Not captured in ATS.",
    }


def build_role_summary(candidates):
    grouped = defaultdict(list)
    for candidate in candidates:
        grouped[candidate["role"]].append(candidate)

    summaries = []
    for role, items in grouped.items():
        office_counts = Counter(item["office_location"] for item in items if item["office_location"] != "Unknown")
        office = office_counts.most_common(1)[0][0] if office_counts else "Not captured in ATS."
        known_scores = [item["interview_score"] for item in items if item["interview_score"] is not None]
        avg_score = round(sum(known_scores) / len(known_scores), 2) if known_scores else None
        known_stages = [item["stage"] for item in items if item["stage"] != "Unknown"]
        stage = max(known_stages, key=stage_rank) if known_stages else "Not captured in ATS."
        statuses = Counter(item["status"] for item in items)
        if statuses.get("Rejected"):
            status = "Need Attention"
        elif any(item["risk_level"] == "High" for item in items):
            status = "Need Attention"
        elif any(item["stage"] == "3rd Interview" for item in items):
            status = "Decision Needed"
        elif any(item["stage"] == "Unknown" for item in items):
            status = "Need Attention"
        else:
            status = "On Track"
        summaries.append(
            {
                "id": slug(role),
                "role": role,
                "office": office,
                "pipeline_count": len(items),
                "current_stage": stage,
                "average_score": avg_score,
                "status": status,
                "status_reason": "Derived from ATS stage, risk, and score mix.",
                "candidate_names": [item["candidate_name"] for item in items],
                "selected_candidate": sorted(
                    items,
                    key=lambda row: (
                        stage_rank(row["stage"]),
                        row["interview_score"] is None,
                        -(row["interview_score"] or 0),
                        row["candidate_name"],
                    ),
                )[0]["candidate_name"],
            }
        )

    summaries.sort(key=lambda row: (row["status"] != "Decision Needed", row["status"] != "Need Attention", row["role"]))
    return summaries


def build_office_summary(candidates):
    grouped = defaultdict(list)
    for candidate in candidates:
        if candidate["office_location"] == "Unknown":
            continue
        grouped[candidate["office_location"]].append(candidate)

    summaries = []
    for office, items in grouped.items():
        known_scores = [item["interview_score"] for item in items if item["interview_score"] is not None]
        avg_score = round(sum(known_scores) / len(known_scores), 2) if known_scores else None
        stage = Counter(item["stage"] for item in items).most_common(1)[0][0] if items else "Not captured in ATS."
        summaries.append(
            {
                "office": office,
                "count": len(items),
                "active": sum(1 for item in items if item["status"] == "Active"),
                "average_score": avg_score,
                "top_stage": stage,
                "status": "Need Attention" if any(item["risk_level"] == "High" for item in items) else "On Track",
            }
        )

    summaries.sort(key=lambda row: (-row["count"], row["office"]))
    return summaries


def build_pipeline_health(candidates):
    stages = Counter(item["stage"] for item in candidates)
    office_counts = Counter(item["office_location"] for item in candidates)
    risk_counts = Counter(item["risk_level"] for item in candidates)
    status_counts = Counter(item["status"] for item in candidates)
    return {
        "stages": stages,
        "offices": office_counts,
        "risks": risk_counts,
        "statuses": status_counts,
    }


def main():
    candidates_raw, funnel_rows = read_workbook()
    candidates = build_candidate_rows(candidates_raw)
    funnel = build_funnel_summary(funnel_rows)

    feedback_rows = parse_feedback_exports(FEEDBACK_CSV, "Contact Name")
    calibration_json = ROOT / "_calibration_rows.json"
    calibration_rows = load_json_rows(calibration_json) if calibration_json.exists() else parse_feedback_exports(CALIBRATION_CSV, "Candidate Name")
    feedback_map = latest_detail(feedback_rows, "Contact Name")
    calibration_map = latest_detail(calibration_rows, "Candidate Name")

    candidate_details = []
    detail_lookup = {}
    for candidate in candidates:
        detail = build_detail_entry(candidate, feedback_map, calibration_map)
        candidate_details.append(detail)
        detail_lookup[candidate["candidate_name"]] = detail

    roles = build_role_summary(candidates)
    offices = build_office_summary(candidates)
    health = build_pipeline_health(candidates)

    total_candidates = len(candidates)
    active_candidates = sum(1 for row in candidates if row["status"] == "Active")
    open_roles = len(roles)
    avg_score_values = [row["interview_score"] for row in candidates if row["interview_score"] is not None]
    avg_score = round(sum(avg_score_values) / len(avg_score_values), 2) if avg_score_values else None
    high_risk = sum(1 for row in candidates if row["risk_level"] == "High")
    need_review = sum(1 for row in candidates if row["stage"] == "Unknown")

    stage_counts = {stage: funnel["stage_counts"].get(stage, 0) for stage in funnel["stage_order"]}

    sorted_candidates = sorted(
        candidates,
        key=lambda row: (
            row["status"] != "Active",
            stage_rank(row["stage"]),
            -(row["interview_score"] or 0) if row["interview_score"] is not None else 99,
            row["candidate_name"],
        ),
    )

    leadership_summary = [
        f"{total_candidates} candidates in the clean ATS export.",
        f"{active_candidates} active records and {need_review} need review.",
        "Unknown stage and location entries are the main cleanup items.",
    ]

    bottlenecks = [
        {"label": "Unknown stage", "value": funnel["stage_counts"].get("Unknown", 0), "note": "Records need manual stage review."},
        {"label": "Unknown office", "value": funnel["office_counts"].get("Unknown", 0), "note": "Location was not captured in the export."},
        {"label": "Rejected", "value": funnel["status_counts"].get("Rejected", 0), "note": "Do not move these forward."},
        {"label": "Withdrawn", "value": funnel["status_counts"].get("Withdrawn", 0), "note": "Inactive candidate record."},
        {"label": "High risk", "value": funnel["risk_counts"].get("High", 0), "note": "Needs leadership review."},
    ]

    data = {
        "meta": {
            "header": "ADAPTOVATE Talent Dashboard",
            "title": "Talent Dashboard",
            "subtitle": "Live pipeline. Interview feedback. Candidate assessment.",
            "period": "Clean workbook import",
            "privacyNote": "This internal dashboard uses ATS interview and feedback data for recruiting management. Candidate names are shown only in pipeline and role assessment views. Emails, CVs, salary details, contracts, and unnecessary personal data are not displayed.",
        },
        "summary": {
            "leadershipSummary": leadership_summary,
            "leadershipInsights": [
                "Toronto is the smallest office queue.",
                "Unknown stage records need manual cleanup.",
                "Most candidates are still active.",
                "A few high-risk records need review.",
            ],
            "kpis": {
                "totalCandidates": total_candidates,
                "activeCandidates": active_candidates,
                "openRoles": open_roles,
                "averageScore": avg_score,
                "highRisk": high_risk,
                "needReview": need_review,
            },
            "officeSummary": offices,
            "roleSummary": roles,
            "stageHealth": stage_counts,
            "bottlenecks": bottlenecks,
        },
        "workbookSummary": {
            "sheetNames": ["Candidates", "Funnel Summary", "Field Sources", "Data Issues"],
            "sourceWorkbook": "talent_dashboard_clean_output.xlsx",
            "sourceFiles": {
                "feedback": "export_feedback.csv",
                "calibration": "export_interview_calibration.csv",
                "workbook": "talent_dashboard_clean_output.xlsx",
            },
        },
        "candidates": candidates,
        "candidateDetails": candidate_details,
        "candidateLookup": detail_lookup,
        "funnel": funnel,
        "rawCounts": {
            "stage": dict(funnel["stage_counts"]),
            "office": dict(funnel["office_counts"]),
            "status": dict(funnel["status_counts"]),
            "risk": dict(funnel["risk_counts"]),
        },
        "selectionDefaults": {
            "office": "Global",
            "roleId": roles[0]["id"] if roles else "",
            "candidateId": sorted_candidates[0]["id"] if sorted_candidates else "",
        },
        "scorecardTemplates": {
            "consulting mindset": "Not captured in ATS.",
            "problem solving": "Not captured in ATS.",
            "client presence": "Not captured in ATS.",
            "adaptability": "Not captured in ATS.",
            "leadership": "Not captured in ATS.",
            "communication": "Not captured in ATS.",
            "commercial awareness": "Not captured in ATS.",
        },
        "pipelineHealth": {
            "stages": dict(health["stages"]),
            "offices": dict(health["offices"]),
            "risks": dict(health["risks"]),
            "statuses": dict(health["statuses"]),
        },
    }

    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "candidates.json").write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "roles.json").write_text(json.dumps(roles, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "funnel.json").write_text(json.dumps(data["funnel"], ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "scorecards.json").write_text(json.dumps(candidate_details, ensure_ascii=False, indent=2), encoding="utf-8")

    OUTPUT_JS.write_text(
        "// Generated from talent_dashboard_clean_output.xlsx and ATS exports.\n"
        "window.TALENT_DATA = "
        + json.dumps(data, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    print(f"Wrote {OUTPUT_JS}")


if __name__ == "__main__":
    main()
